# Importações
import pyautogui as pg
from time import sleep
import openpyxl
import locale
from datetime import datetime, timedelta
import tkinter as tk

# data atual
data_execucao = datetime.now()

primeiro_dia_mes_atual = data_execucao.replace(day=1)

# primeiro dia do mes anterior
primeiro_dia_mes_anterior = primeiro_dia_mes_atual - timedelta(days=1)
primeiro_dia_mes_anterior = primeiro_dia_mes_anterior.replace(day=1)

# ultimo dia do mes anterior
ultimo_dia_mes_anterior = primeiro_dia_mes_atual - timedelta(days=1)

# Formata as datas no formato dd/mm/aaaa
data_inicial_formatada = primeiro_dia_mes_anterior.strftime('%d/%m/%Y')
data_final_formatada = ultimo_dia_mes_anterior.strftime('%d/%m/%Y')

# Metricas
valor_total_inserido = 0
loops_feitos = 0
tempo_inicio = datetime.now()


# Funçao basica para diminuir codigo, pois repete algumas vezes
def tab():
    sleep(0.7)
    pg.press('tab')
    sleep(0.7)


# Funçao para criar o popup
def popup_mensagem(mensagem):
    popup = tk.Tk()
    popup.title("Aviso")
    popup.geometry("300x100")

    texto = tk.Label(popup, text=mensagem, pady=10)
    texto.pack()

    btn_fechar = tk.Button(popup, text="Fechar", command=popup.destroy)
    btn_fechar.pack()

    popup.lift()
    popup.attributes("-topmost", True)

    popup.mainloop()


# Seleciona a localizacao BRASIL // para o padrao da moeda brasileira
locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')

current_date = datetime.now()
competence = (current_date - timedelta(days=current_date.day)).strftime('%m/%Y')

sheet_name = f'CT-es {competence}.xlsx'.replace('/', '-')

# Carrega arquivo excel
workbook = openpyxl.load_workbook(sheet_name)
sheet = workbook.active

# Valor total dos CT-es
vlr_total = sum(row[2] for row in sheet.iter_rows(
    min_row=2, values_only=True) if row[2] is not None)
vlr_total_formatado = locale.currency(vlr_total, grouping=True)
vlr_total_str = locale.format_string('%.*f', (2, vlr_total), grouping=True)

# Observaçoes do CT-e
# Dados omitidos aqui
obs_gerais = f'''DESCRIÇÃO OCULTA'''.upper()

sleep(3)  # Tempo para o usuario trocar para a tela do Hivecloud

# Loop sobre todas as linhas do arquivo excel, apartir da lunha 2 (linha 1 = cabeçalho)
for row in sheet.iter_rows(min_row=2, values_only=True):
    n_doc = row[0]    # Coluna A
    dt_doc = row[1]   # Coluna B
    vlr_doc = row[2]  # Coluna C

    # Data com formato de data python
    data_python = datetime.strptime(str(dt_doc), '%d%m%Y')

    # Formata para dd/mm/aaaa
    data_formatada = data_python.strftime('%d/%m/%Y')

    # Botao adicionar
    pg.click(700, 230)
    sleep(0.7)

    # Tipo do documento
    pg.click(600, 300)
    sleep(0.7)
    pg.write('o')
    sleep(0.7)
    pg.press('enter')
    tab()

    # Descriçao
    descricao = f'DESCRIÇÃO OCULTA'
    pg.write(descricao)
    tab()

    # Numero do documento
    n_doc_str = str(n_doc)
    pg.write(n_doc_str)
    tab()

    # Data do documento
    data_formatada_str = str(data_formatada)
    pg.write(data_formatada_str)
    tab()
    pg.press('tab')  # dois tab's para sair do campo de data (widget calendário)
    sleep(0.7)

    # Valor do documento
    valor_formatado = locale.currency(vlr_doc, grouping=True)
    vlr_doc_str = str(valor_formatado)
    pg.write(vlr_doc_str)
    tab()

    # Salvar
    pg.click(960, 640)
    sleep(0.7)

    valor_total_inserido += vlr_doc
    loops_feitos += 1

# Depois do loop
tempo_execucao = datetime.now() - tempo_inicio

# Valor total do CT-e
pg.click(960, 150)
sleep(0.7)
pg.doubleClick(360, 330)
sleep(0.7)
pg.write(vlr_total_str.replace('.', ''))
sleep(0.7)

# Aba Avançado
pg.click(1060, 150)
sleep(0.7)

# Campo de observaçpes
pg.click(960, 350)
sleep(0.7)
pg.hotkey('ctrl', 'a')
sleep(0.7)
pg.write(obs_gerais)
sleep(0.7)

# Confirmar e Salvar
pg.click(1250, 410)

popup_mensagem(
    f'Tempo de execução: {tempo_execucao}\nQuantidade de loops feitos: {loops_feitos}\nValor total: {vlr_total_str}')
