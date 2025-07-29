import os
import re
try:
    from PyPDF2 import PdfReader
except ImportError:
    print("PyPDF2 não está instalado. Execute: pip install PyPDF2")
    exit(1)
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

def limpar_texto(texto):
    """
    Remove caracteres especiais e limpa o texto para facilitar a extração
    """
    # Remove quebras de linha excessivas e espaços extras
    texto = re.sub(r'\n+', '\n', texto)
    texto = re.sub(r'\s+', ' ', texto)
    # Remove caracteres especiais que podem atrapalhar
    texto = re.sub(r'[^\w\s\+\-\.,/]', ' ', texto)
    return texto.strip()

def extrair_dados_cte(texto):
    """
    Extrai número do contrato, data e valor do frete do texto do PDF
    """
    # Limpa o texto primeiro
    texto_limpo = limpar_texto(texto)
    
    # Padrões múltiplos para extrair os dados (mais robustos)
    padroes_contrato = [
        r'CONTRATO\s+N[º°]\s*(\d+)',
        r'CONTRATO\s+(\d+)',
        r'N[º°]\s*(\d+)',
        r'CONTRATO[\s\n]+N[º°][\s\n]*(\d+)',
        r'CONTRATO[\s\n]+(\d+)'
    ]
    
    padroes_data = [
        r'DATA\s+(\d{2}\/\d{2}\/\d{4})',
        r'DATA[\s\n]+(\d{2}\/\d{2}\/\d{4})',
        r'DATA\s*:?\s*(\d{2}\/\d{2}\/\d{4})',
        r'(?:DATA|Data)\s+(\d{1,2}\/\d{1,2}\/\d{4})'
    ]
    
    padroes_valor_frete = [
        r'Valor frete\s*\+\s*([\d,.]+)',
        r'Valor frete\s*\+?\s*([\d,.]+)',
        r'Valor frete[\s\n]*\+[\s\n]*([\d,.]+)',
        r'frete\s*\+\s*([\d,.]+)',
        r'FRETE\s*\+\s*([\d,.]+)'
    ]
    
    # Extrai os dados usando regex (tenta múltiplos padrões)
    numero_contrato = None
    data = None
    valor_frete = None
    
    # Busca número do contrato
    for padrao in padroes_contrato:
        match = re.search(padrao, texto_limpo, re.IGNORECASE)
        if match:
            numero_contrato = match.group(1)
            break
    
    # Busca data
    for padrao in padroes_data:
        match = re.search(padrao, texto_limpo, re.IGNORECASE)
        if match:
            data_str = match.group(1)
            # Converte para formato DDMMAAAA
            try:
                # Normaliza a data (adiciona zeros se necessário)
                partes = data_str.split('/')
                if len(partes) == 3:
                    dia = partes[0].zfill(2)
                    mes = partes[1].zfill(2)
                    ano = partes[2]
                    data = f"{dia}{mes}{ano}"
                    break
            except:
                continue
    
    # Busca valor do frete
    for padrao in padroes_valor_frete:
        match = re.search(padrao, texto_limpo, re.IGNORECASE)
        if match:
            valor_str = match.group(1)
            # Remove pontos (separadores de milhares) e mantém vírgula como separador decimal
            valor_frete = valor_str.replace('.', '')
            break
    
    return numero_contrato, data, valor_frete

def processar_todos_pdfs(pasta_ctes="CTEs"):
    """
    Processa todos os PDFs na pasta CTEs e extrai os dados
    """
    if not os.path.exists(pasta_ctes):
        print(f"Pasta {pasta_ctes} não encontrada!")
        return []
    
    # Lista todos os arquivos PDF na pasta
    arquivos_pdf = [f for f in os.listdir(pasta_ctes) if f.lower().endswith('.pdf')]
    
    if not arquivos_pdf:
        print("Nenhum arquivo PDF encontrado na pasta CTEs!")
        return []
    
    print(f"Encontrados {len(arquivos_pdf)} arquivos PDF para processar...")
    
    # Dados extraídos
    dados_extraidos = []
    erros = []
    
    for i, nome_arquivo in enumerate(arquivos_pdf, 1):
        caminho_arquivo = os.path.join(pasta_ctes, nome_arquivo)
        
        try:
            print(f"Processando ({i}/{len(arquivos_pdf)}): {nome_arquivo}")
            
            # Lê o PDF
            with open(caminho_arquivo, 'rb') as arquivo:
                leitor_pdf = PdfReader(arquivo)
                
                # Extrai texto de todas as páginas
                texto_completo = ""
                for pagina in leitor_pdf.pages:
                    texto_completo += pagina.extract_text() + "\n"
                
                # Extrai os dados
                numero_contrato, data, valor_frete = extrair_dados_cte(texto_completo)
                
                # Adiciona aos dados extraídos
                dados_extraidos.append({
                    'arquivo': nome_arquivo,
                    'numero_contrato': numero_contrato,
                    'data': data,
                    'valor_frete': valor_frete
                })
                
                print(f"  ✓ Contrato: {numero_contrato}, Data: {data}, Valor: {valor_frete}")
                
        except Exception as e:
            erro_msg = f"Erro ao processar {nome_arquivo}: {str(e)}"
            print(f"  ✗ {erro_msg}")
            erros.append(erro_msg)
    
    if erros:
        print(f"\nErros encontrados ({len(erros)}):")
        for erro in erros:
            print(f"  - {erro}")
    
    return dados_extraidos

def criar_planilha_excel(dados, nome_arquivo="CTes.xlsx"):
    """
    Cria a planilha Excel com os dados extraídos (apenas dados completos)
    """
    # Filtra apenas os dados completos
    dados_completos = [d for d in dados if all([d['numero_contrato'], d['data'], d['valor_frete']])]
    
    if not dados_completos:
        print("Nenhum dado completo encontrado para criar a planilha!")
        return None
    
    # Cria um novo workbook
    wb = Workbook()
    ws = wb.active
    
    if ws is None:
        ws = wb.create_sheet("CTEs")
    else:
        ws.title = "CTEs"
    
    # Cores para formatação
    cor_cabecalho = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Define os cabeçalhos
    headers = ['Nº', 'DATA', 'VALOR']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center')
        cell.fill = cor_cabecalho
    
    # Preenche apenas os dados completos
    for row, dados_item in enumerate(dados_completos, 2):
        ws.cell(row=row, column=1, value=dados_item['numero_contrato'])
        ws.cell(row=row, column=2, value=dados_item['data'])
        ws.cell(row=row, column=3, value=dados_item['valor_frete'])
    
    # Ajusta a largura das colunas
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    
    # Adiciona filtros
    ws.auto_filter.ref = ws.dimensions
    
    # Salva a planilha
    wb.save(nome_arquivo)
    print(f"\nPlanilha salva como: {nome_arquivo}")
    print(f"Dados inseridos na planilha: {len(dados_completos)}")
    
    return nome_arquivo

def gerar_relatorio_estatisticas(dados):
    """
    Gera um relatório com estatísticas dos dados extraídos
    """
    if not dados:
        return
    
    print("\n" + "="*50)
    print("RELATÓRIO DE ESTATÍSTICAS")
    print("="*50)
    
    total_arquivos = len(dados)
    dados_completos = [d for d in dados if all([d['numero_contrato'], d['data'], d['valor_frete']])]
    dados_incompletos = [d for d in dados if not all([d['numero_contrato'], d['data'], d['valor_frete']])]
    
    print(f"Total de arquivos processados: {total_arquivos}")
    print(f"Dados extraídos com sucesso: {len(dados_completos)} ({len(dados_completos)/total_arquivos*100:.1f}%)")
    print(f"Dados incompletos: {len(dados_incompletos)} ({len(dados_incompletos)/total_arquivos*100:.1f}%)")
    
    # Estatísticas por tipo de problema
    sem_contrato = len([d for d in dados_incompletos if not d['numero_contrato']])
    sem_data = len([d for d in dados_incompletos if not d['data']])
    sem_valor = len([d for d in dados_incompletos if not d['valor_frete']])
    
    if dados_incompletos:
        print(f"\nTipos de problemas encontrados:")
        print(f"  - Sem número de contrato: {sem_contrato}")
        print(f"  - Sem data: {sem_data}")
        print(f"  - Sem valor do frete: {sem_valor}")
    
    # Valores estatísticos
    valores_validos = []
    for d in dados_completos:
        try:
            # Converte vírgula para ponto para cálculos
            valor_str = d['valor_frete'].replace(',', '.')
            valor = float(valor_str)
            valores_validos.append(valor)
        except:
            pass
    
    if valores_validos:
        print(f"\nEstatísticas dos valores:")
        # Formata valores com vírgula decimal para exibição
        total = sum(valores_validos)
        media = total / len(valores_validos)
        minimo = min(valores_validos)
        maximo = max(valores_validos)
        
        print(f"  - Valor total: R$ {total:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'))
        print(f"  - Valor médio: R$ {media:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'))
        print(f"  - Valor mínimo: R$ {minimo:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'))
        print(f"  - Valor máximo: R$ {maximo:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'))
    
    # Relatório de arquivos não utilizados
    if dados_incompletos:
        print(f"\nArquivos NÃO UTILIZADOS na planilha ({len(dados_incompletos)}):")
        for item in dados_incompletos:
            motivos = []
            if not item['numero_contrato']:
                motivos.append("sem número do contrato")
            if not item['data']:
                motivos.append("sem data")
            if not item['valor_frete']:
                motivos.append("sem valor do frete")
            
            print(f"  - {item['arquivo']}: {', '.join(motivos)}")

def main():
    """
    Função principal que executa todo o processo
    """
    print("=== EXTRATOR DE DADOS CTEs ===")
    print("Processando arquivos PDF...")
    
    inicio = datetime.now()
    
    # Processa todos os PDFs
    dados = processar_todos_pdfs()
    
    if not dados:
        print("Nenhum dado foi extraído!")
        return
    
    # Cria a planilha Excel
    arquivo_excel = criar_planilha_excel(dados)
    
    # Gera relatório de estatísticas
    gerar_relatorio_estatisticas(dados)
    
    fim = datetime.now()
    tempo_processamento = fim - inicio
    
    print(f"\nTempo de processamento: {tempo_processamento.total_seconds():.2f} segundos")
    
    if arquivo_excel:
        print(f"✅ Processamento concluído com sucesso!")
    else:
        print(f"❌ Nenhuma planilha foi criada - verifique os dados extraídos.")

if __name__ == "__main__":
    main() 