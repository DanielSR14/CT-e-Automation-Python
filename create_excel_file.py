# Importaçoes
import tkinter as tk
from tkinter import filedialog
import os
import glob
import pytesseract
from PIL import Image
from openpyxl import Workbook
from datetime import datetime, timedelta

# define o executavel do tesseract, responsavel pela "leitura" das iamgens
tesseract_path = 'C:\Program Files\Tesseract-OCR'
pytesseract.pytesseract.tesseract_cmd = os.path.join(tesseract_path, 'tesseract.exe')

# define a competencia dos dados coletados
current_date = datetime.now()
competence = (current_date - timedelta(days=current_date.day)).strftime('%m/%Y')



# solicita ao usuario uma pasta contendo as imagens escaneadas
def select_folder():
    folder = filedialog.askdirectory()
    if folder:
        img_process(folder)
        root.destroy()



# processa o texto extraido da imagem, para preencher a planilha

def text_process(text):
    lines = text.split('\n')

    date = ''
    value = ''
    receipt = ''

    # Dados omitidos
    for line in lines:
        if 'Data:' in line:
            date = line.split(': ')[1] 
        elif 'Valor:' in line:
            value = line.split(': ')[1]  
        elif 'Numero:' in line:
            receipt = line.split(': ')[1]

    return date, value, receipt



# verifica se existe um arquivo com o mesmo nome, se nao cria um arquivo excel novo

sheet_name = f'CT-es {competence}.xlsx'.replace('/', '-') # nao pode salvar arquivos com /,
                                                          # por isso o replace
if os.path.isfile(sheet_name):
    pass

else:
    wb = Workbook()
    wb.save(sheet_name)

# processa as imagens, extrai os dados necessarios, e monta um arquivo excel
def img_process(directory):
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'CT-es'


    folder_path = os.path.join(directory, '*.png')

    # encontra todos os arquivos com a extensao .png
    images = glob.glob(folder_path)
    
    # para cada imagem, extrai a data, o valor do documento e o numero do recibo
    for index, image in enumerate(images, start=1):
        text = extract_data(image)
        date, value, receipt = text_process(text)

        # Insere o numero do recibo na coluna A
        sheet.cell(row=index, column=1, value=receipt)
        # Imsere a data na coluna B
        sheet.cell(row=index, column=2, value=date)
        # Insere o valor do documento na coluna C
        sheet.cell(row=index, column=3, value=value)

        # terminal logs 
        print(f'Recibo {index} feito!')

    wb.save(sheet_name)

def extract_data(image):
    text = pytesseract.image_to_string(Image.open(image))
    return text


# Usando TKinter para ficar mais conveniente a escolha de uma pasta para o usuario

root = tk.Tk()
root.geometry("300x100")
root.title('Selecione a pasta com as imagens escaneadas:')

# botão para selecionar a pasta
btn_select = tk.Button(root, text='Selecionar Pasta', command=select_folder)
btn_select.pack(pady=20)

# inicia a interface TKinter
root.mainloop()
